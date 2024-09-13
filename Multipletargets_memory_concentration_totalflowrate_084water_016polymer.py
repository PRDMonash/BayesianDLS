import os
import torch
import numpy as np
import pandas as pd
import random

#Automation of DLS
from datetime import datetime
import serial
import serial.tools.list_ports
from time import *
from time import sleep
from SF10 import SF10
import subprocess
import pyautogui
import pygetwindow as gw 
import glob
import os.path
from pandas import ExcelWriter
from pandas import ExcelFile
import matplotlib.pyplot as plt
import openpyxl

#Self-optimization
from botorch.models import SingleTaskGP, ModelListGP
from gpytorch.mlls.exact_marginal_log_likelihood import ExactMarginalLogLikelihood
#Fitting
from botorch import fit_gpytorch_model
#Acquisition function
from botorch.acquisition.monte_carlo import qExpectedImprovement
from botorch.optim import optimize_acqf


#Initiliaze target size
target_sizelist=[200, 150, 130]
#Define target size tolerance
tolerance=[10, 7.5, 6] 
#set upper and lower boundaries for concentration and total flowrates
bounds = torch.tensor([[0.5,1], [5, 8]])
#Initialize volume fraction
volumefraction_water=[0.84]
#Initialize volume fraction of polymer since this now consists out of THF and polymer volume fraction
volumefraction_polymer_tot=[0.16]
#Initialize beginning concentration
concentration_initial=[5]
#Initialize beginning concentration for flow parameters function as a single numerical value
concentration_initial_flow=5

#Generate initial dataset new
def generate_initial_data(n,target):
    
    #Generate four random volume fractions between 0.5 and 5
    X1_initial = [random.uniform(0.5, 5) for _ in range(n)]
    #Generate four random flowrates between 1 and 8
    X2_initial = [random.randint(1, 8) for _ in range(n)]
    #Calculate according polymer, water and dilution flowrates for pumps based on volumefraction of water, the total flowrate and concentration
    #Element-wise multiplication using list comprehension
    waterflowrate = [x * y for x, y in zip(volumefraction_water, X2_initial)]
    #Element-wise difference using list comprehension
    polymerflowrate=[((w * (c / i)) * v)  for c, i, w, v in zip(X1_initial,concentration_initial,volumefraction_polymer_tot,X2_initial)]
    #Element-wise multiplication using list comprehension while dilution_flowrate=(cinitial/ccurrent-1)*polymer_flowrate since ccurent=cinitial*polymer_flowrate/(polymer_flowrate+dillution_flowrate)
    dilutionflowrate = [((i / n) - 1) * m for i, m, n in zip(concentration_initial,polymerflowrate, X1_initial)]
    #List of initial flow rates
    print(f"your initial water flowrates are: {waterflowrate} and your intital polymer flow rates are: {polymerflowrate}")
     #List of according initialvolume fractions and total flow rates
    print(f"your initial water volume fractions are: {X1_initial} and your intital total flow rates are: {X2_initial}")
    #Initialize the size list
    y_initial=torch.tensor([])
    #Go trough each pair of initial water and polymer flow rates
    for waterflow,polymerflow,dilutionflow in zip(waterflowrate,polymerflowrate,dilutionflowrate):

        #Insert flowrates into a list such that flow_parameters can read it
        waterflow=[waterflow]
        polymerflow=[polymerflow]
        dilutionflow=[dilutionflow]
        #Set repitions to 1 for automated algorithm
        c=[1] 
        #Run the initital experiments with candidates of flowrates for polymer and water pump
        flow_parameters(waterflow,polymerflow,c,dilutionflow)   
        #Measure the mean intensity after each run and return the current size
        micelle=measure_micelle_size() 
        #Append the measured mean intensity to current set
        y_initial=torch.cat([y_initial, torch.tensor([[micelle]], dtype=torch.float64)])
        
    #Reshape X data before concatenating it, first convert it into numpy array and then reshape it
    X1_initial=np.array(X1_initial)
    X2_initial=np.array(X2_initial)
    X1_initial = X1_initial.reshape(-1, 1)   
    X2_initial = X2_initial.reshape(-1, 1) 
    #Concatenate arrays of volume fraction and total flow rate along the second axis
    X_initial = np.concatenate((X1_initial, X2_initial), axis=1)
    #Calculate score based on sizes
    y_initial_score = -abs(target-y_initial) 
    #Extract the closest value to target size from initial sizes
    best_observerd_value=y_initial_score.max().item()
    #Convert numpy arrays to torch tensors and return them
    X_initial_tensor = torch.tensor(X_initial)
    y_initial_tensor = torch.tensor(y_initial)
    y_initial_score_tensor = torch.tensor(y_initial_score)
    return X_initial_tensor, y_initial_score_tensor, y_initial_tensor, best_observerd_value


#Define the model, fit and acquisition function and optimize acquisition function to return next possible candidates of volume fraction and total flow rate
def get_nextpoints(X_initial_tensor,y_initial_tensor,best_observerd_value,n_points=1):
    #Train surrogate model with flowrates and according mean intensities
    single_model= SingleTaskGP(X_initial_tensor,y_initial_tensor) 
    mll= ExactMarginalLogLikelihood(single_model.likelihood, single_model)  
    #Fitting with predefined function of fit_gpytorch_model
    fit_gpytorch_model(mll) 
    #Define the acquisiton function of monte carlo model
    EI = qExpectedImprovement(model=single_model,best_f=best_observerd_value)
    #Generate `npoints=2` candidates jointly using 20 random restarts and 'raw_samples=size of dataset' raw samples
    canidates,_= optimize_acqf(acq_function=EI,bounds=bounds,q=n_points, num_restarts=3,raw_samples=3)
    return canidates


#Extract the most current measured micelle size by mean intensity
def measure_micelle_size():
       
        #Identifying the most recent excel file in the folder
        folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
        file_type = r'\*xlsx'
        files = glob.glob(folder_path + file_type)
        max_file = max(files, key=os.path.getctime)
        df = pd.read_excel(max_file)

        #Extract the most recent measured mean intensity of micelle and recalculate in nm
        measured_micellesize = 1000*(df.iat[22, 2])
        return measured_micellesize

def flow_parameters(a, b, c, d): 

    #Find the COMs that syringe pumps connected to computer
    PortData = serial.tools.list_ports.comports()
    print(PortData)
    for port in PortData:
        print(f"\033[1;31m{port}\033[0m")
               
    #Pumps Assigned 
    PumpPolymer = SF10('COM10', 'PumpPolymer')
    PumpWater = SF10('COM12', 'PumpWater')
    PumpDilution = SF10('COM11','PumpDilution')


    #Making a new list including repetitions required for water flow rates   
    newlist_water = [ratew for item, ratew in zip(c, a) for repnumber in range(item)]
                
    print ("Resulting flow rates of the water pump according to the specified mixing profile are",a) #print only the list of flow rates of water
    sleep(3)
    print("Resulting sequence of the flow rates of the water pump according to the specified mixing profile and no. of repetitions are",newlist_water) #print the list of flow rates of water multiplied by the repetition
    sleep(3)
    
    #Making a new list including repetitions required for polymer flow rates and given concentrations
    newlist_polymer_flow = [ratep for item, ratep in zip(c, b) for repnumber in range(item)]#added
    newlist_polymer_dilution = [ratep for item, ratep in zip(c, d) for repnumber in range(item)]   #added

    print ("Resulting flow rates of the polymer pump at given concentration {} of according to the specified mixing profile are {} ".format(d,b)) #print only the list of flow rates of polymer
    sleep(3)
 
    print("Resulting sequence of the flow rates of the polymer pump according to the specified mixing profile and no. of repetitions are",newlist_polymer_flow) #print the list of flow rates of polymer multiplied by the repetition
    
    print("Resulting sequence of the flow rates of the dilution pump according to the no. of repetitions are",newlist_polymer_dilution) #print the list of flow rates of polymer multiplied by the repetition
    sleep(3)

    for flowrates in zip(newlist_water,newlist_polymer_flow,newlist_polymer_dilution): #added
        print (flowrates)  #Combines two lists for two iterations at the same time
    for flowrates,(wf,pf,cp) in enumerate (zip(newlist_water,newlist_polymer_flow,newlist_polymer_dilution)): #added
        print (flowrates,wf,pf,cp) #Gives an index to the zipped output as in a list 
        
        print("Experiment {} is starting : flow rate of water pump : {} ml/min, polymer pump : {} ml/min at dilution rate of : {} ml/min".format(flowrates+1,wf,pf,cp))

        print("Water will be passed for 1 minute at 1.5 ml/min flow rate")

        #Calculate the given concentration at which the dilution bath is flowing
        conc=concentration_initial_flow*(pf/(pf+cp))

        print("Experiment {} is starting with a concentration of {}".format(flowrates+1,conc))

        #Wash with water for 1 minute 
        PumpWater.start()
        sleep(0.5)
        PumpWater.changeFlowrate(1.5) 
        sleep(60)
        #added
        print(f"Water and polymer wash will be passed for 30 seconds at flow rates, water: {wf} ml/min and polymer: {pf} ml/min at dilution rate of {cp} ml/min")

        #Water, polymer and dilution pump passes for 30 seconds before actual experiment starts
        PumpPolymer.start()
        sleep(0.5)
        PumpWater.changeFlowrate(wf) , PumpPolymer.changeFlowrate(pf), PumpDilution.changeFlowrate(cp)
        sleep(30)

        #DLS cell volume
        V_cell = 0.17 
        
        #Dead volume due to tubing
        V_dead = 1 
        
        #Sleeptime - time for which the pumps run flling the cell and tubing with micelles
        sleeptime = (V_cell + V_dead)*60/(wf+pf+cp) 

        #Water, dilution and polymer pumps running for sleeptime
        print("The two pumps will run for {} seconds to fill the cell with sample; water pump : {} ml/min and polymer pump : {} ml/min and dilution pump: {} ml/min" .format(sleeptime,wf,pf,cp))

        PumpWater.changeFlowrate(wf) , PumpPolymer.changeFlowrate(pf), PumpDilution.changeFlowrate(cp) 
        sleep(sleeptime)

        PumpWater.stop(), PumpPolymer.stop(), PumpDilution.stop()

        print("Experiment {} is finished. DLS analysis will start now".format(flowrates+1))

        #Activate kalliope screen
        sleep(3)
        hwnd = gw.getWindowsWithTitle('apkw') 
        print(hwnd)
        if hwnd != []:
            try:
                hwnd[0].activate()
            except:
                hwnd[0].minimize()
                hwnd[0].maximize()

        #Click on copy prameters button
        sleep(4)
        copy_click = pyautogui.locateCenterOnScreen("copy1.png", confidence=0.5) 
        print (copy_click)
        pyautogui.moveTo(copy_click,duration=2)
        pyautogui.click(copy_click)

        #Erases the current title
        sleep(2)
        pyautogui.hotkey("backspace")

        #Types the title including current exp no, water flow rate and polymer flow rate
        pyautogui.write("Exp {}_water {}_polymer {}_THF {}".format(flowrates+1,wf,pf,cp))

        #Click on start button
        sleep(3)
        start_click = pyautogui.locateCenterOnScreen("start1.png", confidence=0.5) 
        print (start_click)
        pyautogui.moveTo(start_click,duration=2)
        pyautogui.click(start_click)

        #Monitoring for a new file generation in excel files folder 
        path_to_watch = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
        
        print('The path of the folder where your excel data will be saved is', path_to_watch)
        before = dict ([(f, None) for f in os.listdir (path_to_watch)])
        while 1:
            after = dict ([(f, None) for f in os.listdir (path_to_watch)])
            added = [f for f in after if not f in before]
            if added:
                    print("Added: ", ", ".join (added))
                    break
            else:
                    before = after

        #Identifying the most recent excel file in the folder
        folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
        
        file_type = r'\*xlsx'
        files = glob.glob(folder_path + file_type)
        max_file = max(files, key=os.path.getctime)
        
        print('The data will be extracted and saved in', r"S:\Sci-Chem\PRD\DLS\DLS-Data\Lakshani\Three Pumps Control\excel\data library.xlsx" )
        
        #Exporting specific data from exported file to data library excel file
        source_file_path = max_file  
        
        source_workbook = openpyxl.load_workbook(source_file_path)
        source_sheet = source_workbook['Measurement 0']
        

        data_series = [
            source_sheet['B2'].value,
            source_sheet['C7'].value,
            source_sheet['C8'].value,
            source_sheet['C10'].value,
            source_sheet['C15'].value,
            source_sheet['C16'].value,
            source_sheet['C18'].value,
            source_sheet['C19'].value,
            source_sheet['C21'].value,
            source_sheet['C22'].value,
            source_sheet['C24'].value,
            source_sheet['C25'].value,
            source_sheet['C27'].value,
            source_sheet['C28'].value,
            source_sheet['C30'].value,
            source_sheet['C31'].value,
            source_sheet['C33'].value,
            source_sheet['C34'].value,
            source_sheet['C36'].value,
            source_sheet['C37'].value,
            source_sheet['C39'].value,
            source_sheet['C40'].value  
        ]
        #The path to the target Excel file
        target_file_path = r'excel\data output.xlsx'  
       
        target_workbook = openpyxl.load_workbook(target_file_path)
        target_sheet = target_workbook['Sheet1']

        last_row = target_sheet.max_row + 1 
        columns_to_append = ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W' ]

        for cellvalue, column in enumerate(columns_to_append):
            cell_address = f"{column}{last_row}"
            target_sheet[cell_address] = data_series[cellvalue]

        target_workbook.save(target_file_path)

# Objective function defining the function to be optimized
def objective_function(a,b,d,target,y_initial):
    #Set repitions to 1 for automated algorithm
    c=[1] 
    #Run the experiment with suggested candidates of flowrates for polymer and water pump
    flow_parameters(a,b,c,d) 
    #Measure the mean intensity and return the current size
    micelle=measure_micelle_size() 
    #Append the measured mean intensity to current set
    y_initial=torch.cat([y_initial, torch.tensor([[micelle]], dtype=torch.float64)])
    #Calculate the current score based on the most recent measurement and return score as well as whole data set as torch tensor
    score=-abs(target-micelle) 
    return torch.tensor([score]),y_initial

#Initiliaze the loop with training data(total flowrates, concentrations and according sizes) and read out list
X_initial_tensor, y_initial_tensor, y_initial, best_observerd_value=generate_initial_data(n=4,target=target_sizelist[0])
print(f"the initial concentration and totalflowrates are: {X_initial_tensor} and given sizes of: {abs(y_initial)} while the best observed score to target size is {-best_observerd_value}")

#Set optimization counter to 0
i=0

#Loop through all target sizes 
for j in range(len(target_sizelist)):

    #Read current target value
    target = target_sizelist[j]  
    #Read current tolerance
    tolerance_bounday=tolerance[j]
    #Set new_result such that one enters loop for the first iteration(exceeds all tolerances)
    new_result=20  
    
    #The optimization continues until tolerance of target size is reached
    while new_result > tolerance_bounday: 

        #Add to optimization counter after each loop and print it
        i=i+1 
        print(f"Nr. of optimisation loop: {i}")
        #Extract new candidates based on current data set through optimization of acquisition function and print them
        new_canidates= get_nextpoints(X_initial_tensor,y_initial_tensor,best_observerd_value,1)
        print(f"New canidates are: {new_canidates}")
        
        #Extract individual flowrates for pumps 
        concentration = [new_canidates[0, 0].item()]
        totalflowrate = [new_canidates[0, 1].item()]
        #Calculate according polymer, water and dilution flowrates for pumps based on volumefraction of water, the total flowrate and concentration
        #Element-wise multiplication using list comprehension
        waterflowrate = [x * y for x, y in zip(volumefraction_water,totalflowrate)]
        #Element-wise difference using list comprehension while polymer_flowrate=flow_tot*polymer_volume AND polymer_volume_tot=1-water_volume AND polymer_volume_tot=polymer_volume+dilution_volume AND polymer_volume=concentration_current/concentration AND flow_tot=flow_polymer+flow_dilution+flow_water
        polymerflowrate=[((w * (c / i)) * v)  for c, i, w, v in zip(concentration,concentration_initial,volumefraction_polymer_tot,totalflowrate)]
        #Element-wise multiplication using list comprehension while dilution_flowrate=(cinitial/ccurrent-1)*polymer_flowrate since ccurent=cinitial*polymer_flowrate/(polymer_flowrate+dillution_flowrate)
        dilutionflowrate = [((i / n) - 1) * m for i, m, n in zip(concentration_initial,polymerflowrate, concentration)]
      
        #Print new flowrates
        print(f"your new water flowrate is: {waterflowrate} your new polymer flowrate: {polymerflowrate} your new dilution flowrate: {dilutionflowrate}")
        
        #Start the experiment and conclude new score based on flowrates, according sizes and target size
        new_result,y_initial=objective_function(waterflowrate,polymerflowrate,dilutionflowrate,target,y_initial)
        new_result=new_result.unsqueeze(-1)

        #Append the data such that data set grows for each iteration
        X_initial_tensor=torch.cat([X_initial_tensor,new_canidates])
        y_initial_tensor=torch.cat([y_initial_tensor, new_result])

        new_result=abs(new_result.item())
        
        print(f"the current score after the DLS measurement is {abs(new_result)}")
        print(f"The list of used volume fractions and total flowrates: {X_initial_tensor}")
        print(f"This is the list of scores so far: {-y_initial_tensor}")
        print(f"This is the list of sizes so far: {y_initial}")

        best_init_y=y_initial_tensor.max().item()
        print(f"Best point performs this way {-best_init_y}")

    #If target size is reached, repeat the experiment once more for reproducbility reasons
    #Set repitions to 1 for automated algorithm
    c=[1] 
    new_result,y_initial=objective_function(waterflowrate,polymerflowrate,dilutionflowrate,target,y_initial)
    new_result=new_result.unsqueeze(-1)

    #Append repeated micelle size measurement to torch tensor
    y_initial_tensor=torch.cat([y_initial_tensor, new_result])

    #Append repeated flowrates to torch tensor
    X_initial_tensor=torch.cat([X_initial_tensor,new_canidates]) 
    
    print(f"After running the experiment once more with suggested flowrates the score is: {new_result}")
    print(f"The list of all detected sizes so far is: {y_initial}")

    # Check if it's the last element of list of target sizes
    if j == len(target_sizelist) - 1:
        scale = None  # No next value for the last element of target sizes
    else:
        #If there is new target size than rescale size data on new target size to conclude new scores based on previous intensities and new targetr size
        y_initial_tensor=-abs(target_sizelist[j+1]-y_initial) 
        #Print new list of scores based on new target size
        print(f"Your new list of scores based on the new target: {target_sizelist[j+1]} is: {-y_initial_tensor}")